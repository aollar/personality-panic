"""
Personality Panic — data pipeline.
Reads Personality_Panic_Balance_Lock_v3.xlsx (single source of truth for numbers)
and emits assets/data/gamedata.js (window.PP_DATA) with normalized, structured
requirements/effects so the engine never parses free text at runtime.

Anything NOT in the spreadsheet lives in js/assumptions.js (hand-written), not here.
Re-run after any spreadsheet change:  python scripts/build_data.py
"""
import json, os, re
import openpyxl

XLSX = r"C:\Users\aloss\OneDrive\Desktop\Personality Panic\Personality_Panic_Balance_Lock_v3.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "assets", "data", "gamedata.js")

STAT = {
    "Connection": "connection", "Health": "health", "Career": "career",
    "Happiness": "happiness", "Coolness Factor": "coolness", "Coolness": "coolness",
    "Critical Thinking": "critical", "Enlightenment": "enlightenment", "Money": "money",
    "Pet Happiness": "petHappiness", "Pet Health": "petHealth",
}
MAIN_STATS = ["connection", "health", "career", "happiness"]
UPKEEP_STATS = ["coolness", "critical", "enlightenment", "money"]

BUILDING_ID = {
    "Low Cost Housing": "lowCost", "Luxury Apartments": "luxury", "Park": "park",
    "Air One Supermarket": "airOne", "Regret Burger": "regretBurger",
    "Bro Science Gym": "gym", "Bad Decisions Club": "club",
    "Re-Education Temple": "temple", "High IQ University": "university",
    "Corporate Soul Exchange": "soulExchange", "Debtstreet Capital": "debtstreet",
    "Emotional Baggage Airport": "airport", "Ethical Pet Shop": "petShop", "Mall": "mall",
}

# Archetype cards (art is authoritative: codes read off the card sheets)
ARCHETYPES = {
    "ESTP": {"name": "The Daredevil", "card": "card_00.jpg", "tag": "What could go wrong? Don't answer that."},
    "INFJ": {"name": "The Oracle", "card": "card_01.jpg", "tag": "Booked, blessed, and quietly overwhelmed."},
    "ENTJ": {"name": "The Overlord", "card": "card_02.jpg", "tag": "A five-year plan for your weekend."},
    "ISFP": {"name": "The Vibe Curator", "card": "card_03.jpg", "tag": "Aesthetic first. Consequences later."},
    "ESFP": {"name": "The Showstopper", "card": "card_04.jpg", "tag": "If it's extra, it's correct."},
    "ENFJ": {"name": "The Mentor", "card": "card_05.jpg", "tag": "Part pep talk, part life strategy."},
    "ISTJ": {"name": "The Auditor", "card": "card_06.jpg", "tag": "Calm, correct, and fully itemized."},
    "ISTP": {"name": "The Fixer", "card": "card_07.jpg", "tag": "If it works, it works."},
    "ENTP": {"name": "The Provocateur", "card": "card_08.jpg", "tag": "I'm not arguing. I'm upgrading your side."},
    "ISFJ": {"name": "The Guardian", "card": "card_09.jpg", "tag": "Keeping everyone alive and on time."},
    "ESTJ": {"name": "The Taskmaster", "card": "card_10.jpg", "tag": "Fun has been allocated 3:15pm."},
    "INFP": {"name": "The Dreamer", "card": "card_11.jpg", "tag": "Emotionally booked."},
    "INTP": {"name": "The Architect", "card": "card_12.jpg", "tag": "Overthinking is thinking, twice."},
    "ENFP": {"name": "The Campaigner", "card": "card_13.jpg", "tag": "What if... everything?"},
    "INTJ": {"name": "The Strategist", "card": "card_14.jpg", "tag": "Just a simple plan. 47 steps."},
    "ESFJ": {"name": "The Caregiver", "card": "card_15.jpg", "tag": "Your hype person."},
}

# ---------------------------------------------------------------------------
# Requirement normalization: free text -> structured predicate list (AND).
# Leaf kinds the engine understands:
#   housedLow / housedLux / housedAny / homeless / notHomeless
#   ownsItem:<Item Name> / hasPet / noPet / petFoodAvailable / furnitureOwned
#   fridge / stove / rentDue / hasJob / jobInBuilding / benefitsUnlocked
#   statGte:{stat,pctT} / degree:<Undergrad|Masters|PhD> / degreeProgress:<n>
#   myCamp (temple unlock) / foodSupply
# "Money available" is implicit: engine always checks money >= cost.
def norm_req(action_id, building, text):
    t = (text or "").strip()
    reqs = []
    tl = t.lower()
    def add(kind, **kw): reqs.append(dict(kind=kind, **kw))

    if "must be housed here" in tl: add("housedLow")
    if tl.startswith("luxury apartment") or "luxury apartment +" in tl or tl == "luxury apartment":
        add("housedLux")
    if "own a pet" in tl or "own pet" in tl or tl.startswith("owns pet"):
        if "or volunteer" not in tl: add("hasPet")
    if "no pet owned" in tl: add("noPet")
    if "pet food" in tl: add("petFoodAvailable")
    if "furniture owned" in tl: add("furnitureOwned")
    if "homeless" in tl and "rent cycle" in tl: add("homeless"); add("rentDue")
    elif tl == "homeless": add("homeless")
    if tl == "rent cycle": add("rentDue")
    if "rent cycle + luxury" in tl: add("rentDue"); add("housedLux")
    if "requires fridge" in tl or tl == "fridge": add("fridge")
    if "fridge + stove" in tl: add("fridge"); add("stove")
    if "dressy clothes" in tl: add("ownsItem", item="Dressy Clothes")
    if "job assigned at corporate soul exchange" in tl: add("jobInBuilding")
    if tl.startswith("current job"): add("hasJob")
    if "benefits unlocked" in tl: add("benefitsUnlocked")
    if "computer recommended" in tl: pass  # recommendation only
    # thresholds (VALUES ARE ASSUMPTIONS — see js/assumptions.js)
    if "coolness threshold" in tl: add("statGte", stat="coolness", pctT=0.40)
    if "money threshold" in tl: add("statGte", stat="money", pctT=0.50)
    if "enlightenment threshold" in tl: add("statGte", stat="enlightenment", pctT=0.50)
    if "critical thinking threshold" in tl and action_id != "A084":
        add("statGte", stat="critical", pctT=0.30)
    if action_id == "A084":  # Ask for Promotion: must meet next-tier job reqs (engine-side)
        add("promotionEligible")
    if action_id == "A070": add("degreeProgress", n=3)
    if action_id == "A071": add("degree", degree="Undergrad"); add("degreeProgress", n=6)
    if action_id == "A072": add("degree", degree="Masters"); add("degreeProgress", n=10)
    if action_id in ("A062", "A063"): add("myCamp")   # advanced temple actions gated by My Camp
    if action_id == "A058":  # Buy My Camp: one copy per lifetime of doubt (Austin 2026-07-09)
        add("notFlag", flag="myCamp", msg="You already own My Camp")
    if action_id in ("A026", "A027"):  # ALL groceries need a fridge at home, not just 4-week
        add("fridge")                  # (Austin 2026-07-09; sheet only gated A028)
    if action_id == "A076": pass  # Get/Change Job: dialog filters individual jobs
    return reqs

# Effect normalization from the Unlock/Effect + notes columns.
def norm_fx(action_id, building, unlock, category, name):
    u = (unlock or "").lower(); fx = []
    def add(kind, **kw): fx.append(dict(kind=kind, **kw))
    if "fills hunger" in u or category == "Food" and building == "Regret Burger": add("eat")
    if "fills hunger" not in u and name in ("Buy $22 Smoothie",): pass
    if "counts as relaxing" in u or "counts as major relaxation" in u or "counts as mini-relax" in u:
        add("relax")
    if "reduces stress" in u: add("relax")
    if "ends hunger penalty" in u: pass
    if "adds 1 week food supply" in u: add("foodSupply", weeks=1)
    if "adds premium food supply" in u: add("foodSupply", weeks=1, premium=True)
    if "adds 4 weeks food supply" in u: add("foodSupply", weeks=4)
    if "better food efficiency at home" in u: add("foodSupply", weeks=2, premium=True)
    if "unlocks pet system" in u: add("adoptPet")
    if "prevents pet hunger warning" in u: add("petFood", feedings=4)  # feedings = ASSUMPTION
    if "passive pet happiness" in u: add("petToy")
    if "degree progress" in u: add("degreeProgress")
    if "unlocks mid-tier jobs" in u: add("grantDegree", degree="Undergrad")
    if "unlocks high-tier jobs" in u: add("grantDegree", degree="Masters")
    if "unlocks elite jobs" in u: add("grantDegree", degree="PhD")
    if "selects current job tier" in u: add("openJobDialog")
    if "upgrade job tier" in u: add("promote")
    if "lose job tier until rehired" in u: add("quitJob")
    if "keeps player housed" in u: add("payRent", tier="low")
    if "keeps luxury apartment" in u: add("payRent", tier="lux")
    if "enough for casual clothes" in u: add("supportCheque")
    if "survive turn" in u: add("sleepRough")
    if "unlocks advanced temple actions" in u: add("unlock", flag="myCamp")
    if "low risk" in u: add("invest", risk="low")
    if "medium risk" in u: add("invest", risk="medium")
    if "high risk" in u: add("invest", risk="high")
    if "reduces bad finance event chance" in u: add("unlock", flag="tinyPrint")
    if "future debt penalty" in u: add("loan")
    if "reduces bad travel event chance" in u: add("unlock", flag="travelInsurance")
    if "see items_mall" in u:
        cat = {"A112": "Transportation", "A113": "Electronics/Appliances",
               "A114": "Furniture", "A115": "Clothing"}[action_id]
        add("openShop", group=cat)
    if "avoids bad pet shop consequences" in u or "possible pet shop penalty" in u: pass  # flavor for MVP
    if "uses owned furniture bonus" in u: pass
    return fx

def num(v):
    if v is None or v == "": return 0.0
    try: return float(v)
    except (TypeError, ValueError): return 0.0

# ---------------------------------------------------------------------------
# Weekend Update card system (v3 Cards sheet).
# Status cards are ENGINE-TRIGGERED (their rows here supply display text);
# investment cards resolve per held asset per turn; exactly 1 event card is
# drawn per turn using the standing weights below. Redraw on unmet requirement.
# ---------------------------------------------------------------------------
CARD_REQ = {
    "none": [],
    "owns car": [{"kind": "ownsItem", "item": "Car"}],
    "owns bicycle": [{"kind": "ownsItem", "item": "Bicycle"}],
    "owns fridge": [{"kind": "ownsItem", "item": "Fridge"}],
    "owns any tech item": [{"kind": "ownsAnyTech"}],
    "owns any tech or appliance": [{"kind": "ownsAnyTechOrAppliance"}],
    "employed": [{"kind": "hasJob"}],
    "not homeless": [{"kind": "notHomeless"}],
    "has living pet": [{"kind": "hasPet"}],
    "has living pet + any furniture or tech": [{"kind": "hasPet"}, {"kind": "ownsAnyFurnOrTech"}],
    "ate at regret burger last turn": [{"kind": "prevAteRegret"}],
    "visited bro science gym last turn": [{"kind": "prevGym"}],
}
# Effects the free text under-specifies, keyed by card ID (parsed stats still apply):
CARD_FX = {
    "E07": [{"kind": "forceWalk"}],          # transport counts as Walking this turn
    "E11": [{"kind": "clearFood"}],          # stored groceries spoiled
    "E25": [{"kind": "rentMod", "mult": 1.25}],
    "E26": [{"kind": "rentMod", "mult": 0.5}],
}
STAT_RE = re.compile(
    r"(Money|Health|Career|Connection|Happiness|Coolness|Critical Thinking|Enlightenment"
    r"|Pet Happiness|Pet Health)\s*(\+/-|[+-])\s*\w+\s*\((\d+(?:\.\d+)?)%T\)")

def parse_weekend(wb):
    ws = wb["Cards"]
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    cards, weights, invest_odds = [], {}, {}
    STAND_KEY = {"last place": "last", "2nd / 3rd place": "mid", "2nd / 3rd": "mid", "1st place": "first"}
    in_list = False
    for r in rows:
        c0 = str(r[0]).strip() if r[0] is not None else ""
        # standing weights block: Last/2nd/1st rows follow the header row
        if c0.lower() in STAND_KEY and not in_list and r[1] is not None and num(r[1]) <= 1:
            if c0.lower() in ("last place", "2nd / 3rd place", "1st place"):
                weights[STAND_KEY[c0.lower()]] = {
                    "majPos": num(r[1]), "minPos": num(r[2]),
                    "minNeg": num(r[3]), "majNeg": num(r[4])}
                continue
        # investment odds block: Asset | Standing | BG SG SL BL
        if c0 in ("Crypto", "Stocks") and r[1] is not None:
            key = str(r[1]).strip().lower()
            invest_odds.setdefault(c0.lower(), {})[STAND_KEY.get(key, key)] = [
                num(r[2]), num(r[3]), num(r[4]), num(r[5])]
            continue
        if c0 in ("Bonds", "Savings"):
            invest_odds[c0.lower()] = "safe"
            continue
        if c0 == "ID":
            in_list = True
            continue
        if in_list and re.match(r"^[SIE]\d\d$", c0):
            _id, name, typ, pol, mag, req, eff, flav = [
                ("" if v is None else str(v).strip()) for v in r[:8]]
            stats = []
            for m in STAT_RE.finditer(eff):
                pct = float(m.group(3)) / 100.0
                sign = m.group(2)
                if sign == "+/-": sign = "+"   # swing cards: magnitude only, sign from odds
                stats.append({"stat": STAT[m.group(1)], "pct": pct if sign == "+" else -pct})
            card = {
                "id": _id, "name": name, "type": typ.lower(),
                "polarity": pol.lower(), "magnitude": mag.lower() if mag and mag != "—" else "",
                "req": CARD_REQ.get(req.lower(), []) if typ == "Event" else [],
                "stats": stats, "fx": CARD_FX.get(_id, []),
                "flavor": flav, "effectText": eff,
            }
            if typ == "Event":
                cls = ("maj" if card["magnitude"] == "major" else "min") + \
                      ("Pos" if card["polarity"] == "positive" else "Neg")
                card["cls"] = cls
            cards.append(card)
    # Which card face shows for each investment outcome (from the sheet's
    # "Card IDs used" column: crypto I01/I03/I02, stocks I04/I06/I05; stocks
    # have no big-loss card so I05's -Standard is the worst stocks can do).
    by_id = {c["id"]: c for c in cards}
    def mag(cid, sign=1):
        pcts = [s["pct"] for s in by_id[cid]["stats"] if s["stat"] == "money"]
        return abs(pcts[0]) * sign
    invest_fx = {
        "crypto":  {"bigGain": ["I01", mag("I01")], "smallGain": ["I03", mag("I03")],
                    "smallLoss": ["I03", -mag("I03")], "bigLoss": ["I02", -mag("I02")]},
        "stocks":  {"bigGain": ["I04", mag("I04")], "smallGain": ["I06", mag("I06")],
                    "smallLoss": ["I05", -mag("I05")], "bigLoss": ["I05", -mag("I05")]},
        "bonds":   {"pay": ["I07", mag("I07")]},
        "savings": {"pay": ["I08", mag("I08")]},
    }
    # Upkeep time penalties (Settings sheet, "UPKEEP TIME PENALTIES" block)
    hunger, stress = 4, 2
    for r in wb["Settings"].iter_rows(values_only=True):
        label = str(r[0]) if r[0] else ""
        if label.startswith("Hunger"): hunger = int(num(r[1]) * -1 if num(r[1]) < 0 else num(r[1]))
        if label.startswith("Stress"): stress = int(num(r[1]) * -1 if num(r[1]) < 0 else num(r[1]))
    assert len([c for c in cards if c["type"] == "event"]) == 30, "expected 30 event cards"
    assert len(weights) == 3 and "crypto" in invest_odds, "weights/odds blocks not parsed"
    return {
        "statusTu": {"hunger": hunger, "stress": stress, "minTu": 1},
        "weights": weights, "investOdds": invest_odds, "investFx": invest_fx,
        "cards": cards,
    }

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ---- Personalities ----
    ws = wb["Personalities"]
    personalities = {}
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]: continue
        code = str(row[0]).strip()
        personalities[code] = {
            "code": code,
            "mainStrength": STAT[row[1].strip()], "upkeepStrength": STAT[row[2].strip()],
            "mainWeakness": STAT[row[3].strip()], "upkeepWeakness": STAT[row[4].strip()],
            "engine": bool(row[5] and "Engine" in str(row[5])),
            **ARCHETYPES[code],
        }

    # ---- Settings / modifiers ----
    mods = {  # from Settings sheet, PERSONALITY MODIFIERS v2 block
        "connection": {"strength": 0.20, "weakness": -0.10, "cap": 1.45},
        "health":     {"strength": 0.20, "weakness": -0.10, "cap": 1.45},
        "happiness":  {"strength": 0.20, "weakness": -0.10, "cap": 1.45},
        "career":     {"strength": 0.10, "weakness": -0.10, "cap": 1.25},
        "coolness":   {"strength": 0.15, "weakness": -0.08, "cap": 1.35},
        "critical":   {"strength": 0.15, "weakness": -0.08, "cap": 1.35},
        "enlightenment": {"strength": 0.15, "weakness": -0.08, "cap": 1.35},
        "money":      {"strength": 0.10, "weakness": -0.08, "cap": 1.25},
    }
    settings = {
        "gameLengths": {"short": 100, "medium": 500, "long": 1000},
        "timeUnitsPerTurn": 40,
        "baseTimeUnits": 40,  # costs used AS AUTHORED (1-3 TU): 40 TU/turn = many actions per day (Austin 2026-07-06)
        "rentIntervalTurns": 4,
        "turnTimerOptions": [30, 60, 90, 120, 0],  # 0 = Unlimited
        "maxPlayers": 4, "minParticipants": 2,
        "incomeMultiplierCap": 1.25,
        "modifiers": mods,
        "mainStats": MAIN_STATS, "upkeepStats": UPKEEP_STATS,
    }

    # ---- Actions_Master ----
    ws = wb["Actions_Master"]
    actions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).startswith("A"): continue
        (aid, bld, name, cat, req, tu, costPct, *_rest) = row[:7] + (None,)
        r = row
        gains = []
        if r[10]: gains.append({"stat": STAT[str(r[10]).strip()], "pct": num(r[11])})
        if r[15]: gains.append({"stat": STAT[str(r[15]).strip()], "pct": num(r[16])})
        petGain = {"stat": STAT[str(r[20]).strip()], "pct": num(r[21])} if r[20] else None
        pens = []
        petGains = [petGain] if petGain else []
        if r[22]: pens.append({"stat": STAT[str(r[22]).strip()], "pct": num(r[23])})
        if r[24] and str(r[24]).strip() in STAT:
            s2 = STAT[str(r[24]).strip()]
            if s2.startswith("petH"):  # sheet quirk (A106): 2nd pet GAIN parked in penalty column
                petGains.append({"stat": s2, "pct": num(r[25])})
            else:
                pens.append({"stat": s2, "pct": num(r[25])})
        # per-ID overrides where free text under-specifies
        aid_s = str(aid)
        reqs = norm_req(aid_s, str(bld).strip(), str(req or ""))
        if aid_s == "A015":  # Work From Home: "computer + desk + job" part of the text
            reqs += [{"kind": "ownsItem", "item": "Computer"}, {"kind": "ownsItem", "item": "Desk"}, {"kind": "hasJob"}]
        if aid_s == "A077":  # CSE Work still requires your job to BE at CSE
            reqs = [{"kind": "jobInBuilding"}]
        if aid_s == "A024":  # support cheque: fx pays it flat; drop row gain to avoid double pay
            gains = []
        fx = norm_fx(aid_s, str(bld).strip(), str(r[26] or ""), str(cat or "").strip(), str(name).strip())
        if aid_s in ("A007", "A105"):  # feeding must mark the pet fed (A007 also consumes pet food)
            fx.append({"kind": "feedPet"})
        actions.append({
            "id": aid_s, "building": BUILDING_ID[str(bld).strip()], "name": str(name).strip(),
            "category": str(cat or "").strip(), "tu": int(num(tu)), "costPct": num(costPct),
            "gains": gains, "petGains": petGains, "penalties": pens,
            "req": reqs,
            "fx": fx,
            "note": str(r[26] or "").strip(),
        })

    # ---- Items_Mall ----
    ws = wb["Items_Mall"]
    items = []
    OUTFIT_ORDER = {"Casual Clothes": 1, "Dressy Clothes": 1, "Smart Clothes": 2, "Business Clothes": 3}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not str(row[0]).startswith("I"): continue
        req = str(row[4] or "").strip().lower()
        reqs = []
        if "luxury apartment" in req and "recommended" not in req: reqs.append({"kind": "housedLux"})
        if req == "housing": reqs.append({"kind": "notHomeless"})
        if "own pet" in req: reqs.append({"kind": "hasPet"})
        if "casual clothes" == req: reqs.append({"kind": "ownsItem", "item": "Casual Clothes"})
        if "smart clothes" == req: reqs.append({"kind": "ownsItem", "item": "Smart Clothes"})
        items.append({
            "id": str(row[0]), "group": str(row[1]).strip(), "name": str(row[2]).strip(),
            "slot": str(row[3]).strip(), "req": reqs, "costPct": num(row[5]),
            "bonus": ({"stat": STAT[str(row[9]).strip()], "pct": num(row[10])} if row[9] else None),
            "penalty": ({"stat": STAT[str(row[11]).strip()], "pct": num(row[12])} if row[11] else None),
            "effect": str(row[13] or "").strip(),
            "outfitRank": OUTFIT_ORDER.get(str(row[2]).strip(), 0),
        })

    # ---- Jobs_Named ----
    ws = wb["Jobs_Named"]
    jobs = []
    fxre = re.compile(r"([+-]\d+)\s+([A-Za-z ]+?)(?:,|$| bonus)")
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0] or str(row[0]).startswith("CANONICAL"): continue
        bld = str(row[0]).strip()
        if bld not in BUILDING_ID: continue
        effects = []
        for m in fxre.finditer(str(row[5] or "")):
            sname = m.group(2).strip()
            if sname in STAT: effects.append({"stat": STAT[sname], "amtT100": int(m.group(1))})
        reqtext = str(row[6] or "")
        jreq = {"clothes": None, "computer": "Computer" in reqtext, "degree": None, "stats": []}
        for c in ("Business Clothes", "Smart Clothes", "Dressy Clothes", "Casual Clothes"):
            if c in reqtext: jreq["clothes"] = c; break
        for d, key in (("Undergrad", "Undergrad"), ("Master's", "Masters"), ("PhD", "PhD")):
            if d in reqtext: jreq["degree"] = key
        for m in re.finditer(r"([A-Za-z ]+?)\s+(\d+)\+", reqtext):
            sname = m.group(1).strip()
            if sname in STAT: jreq["stats"].append({"stat": STAT[sname], "pctT": int(m.group(2)) / 100.0})
        jobs.append({
            "building": BUILDING_ID[bld], "tier": str(row[1]).strip(), "name": str(row[2]).strip(),
            "basePayT100": num(row[3]), "careerGainT100": num(row[4]),
            "effects": effects, "req": jreq, "reqText": reqtext.strip(),
        })

    # ---- Pets ----
    ws = wb["Pets"]
    pets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[0] or "")
        if not t.endswith(" Pet"): continue
        code = t.replace(" Pet", "").strip()
        pets[code] = {"main": STAT[str(row[1]).strip()], "upkeep": STAT[str(row[2]).strip()]}

    # ---- Audio ----
    music = {
        "overmap": "PP-Overmap.mp3", "rentDue": "PP-OverMapRentisDue.mp3",
        "airOne": "PP-AirOneSuperMarket.mp3", "soulExchange": "PP-CorporateSoulExchangeOffice.mp3",
        "university": "PP-HighIQUniversity.mp3", "petShop": "PP-EthicalPetShop.mp3",
        "regretBurger": "PP-RegretBurger.mp3", "gym": "PP-BroScienceGym.mp3",
        "park": "PP-Park.mp3", "airport": "PP-EmotionalBaggageAirport.mp3",
        "debtstreet": "PP-DebtStreetCapital.mp3", "club": "PP-BadDecisionsClub.mp3",
        "luxury": "PP-LuxuryApartment.mp3", "lowCost": "PP-LowCostHousing.mp3",
        "mall": "PP-Mall.mp3", "temple": "PP-Temple.mp3",  # track delivered 2026-07 (gap closed)
    }
    sfx = {
        "walk": "PP-Walking.wav", "bike": "PP-BicycleBell.mp3", "car": "PP-CarAcceleration.mp3",
        "click": "PP-MenuButtonClicked.wav", "eat": "PP-EatingFastFood.mp3", "money": "PP-Money.mp3",
    }

    # ---- Buildings + map coords (re-traced 2026-07-19 for Austin's NEW map art,
    # "Game Map (12)": full city redesign — central park roundabout + bottom
    # boulevard + left S-road. Baked fake HUD occupies x 0-19 (real HUD covers it).
    # pos      = building center (map hotspot fallback / labels)
    # entrance = where the DOOR meets the pavement — the walker stands/arrives here
    # doors    = road node(s) the entrance connects to
    buildings = {
        "lowCost": {"name": "Low Cost Housing", "scene": "low_cost_housing.jpg", "pos": [44.5, 72.0],
                    "entrance": [45.4, 81.1], "doors": ["lowFront"]},
        "luxury": {"name": "Luxury Apartments", "scene": "luxury_apartments.jpg", "pos": [86.5, 15.0],
                   "entrance": [84, 34.8], "doors": ["n7"]},
        # the park is the roundabout's island — open entry from any ring stretch
        "park": {"name": "Almost Fine Park", "scene": "park.jpg", "pos": [55.0, 40.0],
                 "entrance": [55.5, 48.0], "doors": ["n3", "rE2", "rNE2", "rNW1b", "rS1", "rS2", "rSE1", "rSW1", "rTop", "rW2"]},
        "airOne": {"name": "Air One Supermarket", "scene": "air_one_supermarket.jpg", "pos": [22.3, 33.0],
                   "entrance": [24.6, 47.5], "doors": ["n1"]},
        "regretBurger": {"name": "Regret Burger", "scene": "regret_burger.jpg", "pos": [69.0, 73.0],
                         "entrance": [70.4, 84.6], "doors": ["burgerFront"]},
        "gym": {"name": "Bro Science Gym", "scene": "bro_science_gym.jpg", "pos": [28.5, 77.0],
                "entrance": [29.9, 92.9], "doors": ["bW1"]},
        "club": {"name": "Bad Decisions Club", "scene": None, "video": "bdc_scene.mp4", "pos": [47.5, 13.0],
                 "entrance": [49.4, 23.6], "doors": ["clubFront"]},
        "temple": {"name": "Re-Education Temple", "scene": "re_education_temple.jpg", "pos": [26.0, 52.0],
                   "entrance": [28.7, 67.8], "doors": ["n2"]},
        "university": {"name": "High IQ University", "scene": "high_iq_university.jpg", "pos": [28.5, 15.0],
                       "entrance": [32, 28.9], "doors": ["uniFront"]},
        "soulExchange": {"name": "Corporate Soul Exchange", "scene": "corporate_soul_exchange.jpg", "pos": [91.5, 43.0],
                         "entrance": [91.3, 63.6], "doors": ["n17"]},
        "debtstreet": {"name": "Debtstreet Capital", "scene": "debtstreet_capital.jpg", "pos": [74.3, 38.0],
                       "entrance": [76.8, 52.9], "doors": ["debtFront"]},
        "airport": {"name": "Emotional Baggage Airport", "scene": "emotional_baggage_airport.jpg", "pos": [64.5, 12.0],
                    "entrance": [64.5, 25.1], "doors": ["rNE2", "rTop"]},
        "petShop": {"name": "Ethical Pet Shop", "scene": "ethical_pet_shop.jpg", "pos": [82.5, 76.0],
                    "entrance": [83, 88.1], "doors": ["n13"]},
        "mall": {"name": "Mall", "scene": "mall.jpg", "pos": [56.0, 72.0],
                 "entrance": [56.5, 82.1], "doors": ["n9"]},
    }

    # Road graph for the walking avatar (percent coords on the 1672x941 map).
    # NEW ART: one ring road circles Almost Fine Park; a boulevard runs along the
    # bottom; the left side is an S-road university -> Air One -> temple -> gym.
    roadNodes = {
        "rTop": [56.6, 26.3],
        "rNE2": [68.2, 31.5],
        "rE1": [71.5, 36],
        "rE2": [72, 42],
        "rSE1": [70.5, 48],
        "rSE2": [67.8, 56.4],
        "rS1": [62.4, 61.7],
        "rS2": [54.8, 63.3],
        "rSW1": [47.5, 61.5],
        "rSW2": [43.1, 60.2],
        "rW1": [38.2, 50.2],
        "rW2": [38.4, 41.3],
        "rNW1": [40.7, 35.8],
        "rNW1b": [43.3, 31.5],
        "clubFront": [51.2, 27.8],
        "uniFront": [34.1, 35.5],
        "leftRd1": [33.6, 41.7],
        "swC1": [39.6, 63.5],
        "swC2": [38, 70],
        "swC3": [37.5, 80],
        "bW1": [34.9, 94.8],
        "bJ1": [38.5, 88],
        "lowFront": [44, 87.5],
        "b2": [50, 92],
        "b3": [62.6, 92.5],
        "burgerFront": [70.1, 90.4],
        "b4": [76.6, 92.6],
        "b5": [88.8, 93],
        "bE": [93.7, 89.8],
        "eR1": [95.1, 78.2],
        "eR2": [92.7, 73.4],
        "sW1": [80.6, 63.9],
        "mE1": [74, 70],
        "debtFront": [74.9, 58.7],
        "dLux1": [74.7, 32.4],
        "n1": [30.8, 46.6],
        "n2": [32.9, 72.8],
        "n3": [63.3, 28.2],
        "n4": [35.5, 50.5],
        "n6": [38.5, 56.9],
        "n7": [79.8, 34.2],
        "n9": [56.1, 88.3],
        "n13": [83, 92.4],
        "n14": [71.4, 56.5],
        "n17": [89.9, 69.5],
    }
    roadEdges = [
        ["rNE2", "rE1"],
        ["rE1", "rE2"],
        ["rE2", "rSE1"],
        ["rSE1", "rSE2"],
        ["rSE2", "rS1"],
        ["rS1", "rS2"],
        ["rS2", "rSW1"],
        ["rSW1", "rSW2"],
        ["rSW2", "rW1"],
        ["rW1", "rW2"],
        ["rW2", "rNW1"],
        ["rNW1", "rNW1b"],
        ["uniFront", "leftRd1"],
        ["rSW2", "swC1"],
        ["swC1", "swC2"],
        ["swC2", "swC3"],
        ["swC3", "bJ1"],
        ["bW1", "bJ1"],
        ["b5", "bE"],
        ["bE", "eR1"],
        ["eR1", "eR2"],
        ["debtFront", "sW1"],
        ["rNE2", "dLux1"],
        ["uniFront", "rW2"],
        ["n1", "leftRd1"],
        ["swC2", "n2"],
        ["rNW1b", "clubFront"],
        ["clubFront", "rTop"],
        ["n3", "rTop"],
        ["n4", "n1"],
        ["n4", "n6"],
        ["n6", "swC1"],
        ["n1", "rW1"],
        ["bJ1", "lowFront"],
        ["lowFront", "b2"],
        ["b2", "n9"],
        ["debtFront", "n14"],
        ["n14", "rSE1"],
        ["rSE2", "n14"],
        ["bJ1", "b2"],
        ["n9", "b3"],
        ["b3", "b2"],
        ["b3", "burgerFront"],
        ["b4", "burgerFront"],
        ["b3", "b4"],
        ["b4", "n13"],
        ["n13", "b5"],
        ["rE1", "dLux1"],
        ["n17", "eR2"],
        ["n17", "sW1"],
        ["uniFront", "rW1"],
        ["n4", "uniFront"],
        ["n6", "rW1"],
        ["n1", "rW2"],
        ["n3", "rNE2"],
        ["n7", "dLux1"],
    ]

    data = {
        "settings": settings, "personalities": personalities, "actions": actions,
        "items": items, "jobs": jobs, "pets": pets, "music": music, "sfx": sfx,
        "buildings": buildings, "roadNodes": roadNodes, "roadEdges": roadEdges,
        "weekend": parse_weekend(wb),
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// GENERATED by scripts/build_data.py from Personality_Panic_Balance_Lock_v3.xlsx\n")
        f.write("// Do not hand-edit numbers here; edit the spreadsheet and re-run the script.\n")
        f.write("var PP_DATA = ")
        f.write(json.dumps(data, indent=1))
        f.write(";\nif (typeof window !== 'undefined') window.PP_DATA = PP_DATA;\n")
        f.write("if (typeof module !== 'undefined') module.exports = PP_DATA;\n")
    print("actions:", len(actions), "items:", len(items), "jobs:", len(jobs),
          "pets:", len(pets), "personalities:", len(personalities))
    print("wrote", os.path.abspath(OUT))

if __name__ == "__main__":
    main()
